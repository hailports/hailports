#!/usr/bin/env python3
"""Apply the 0611 evidence-based reprioritization into Craig's live workbook copy (all 3 tabs)."""
import re
import openpyxl

LIVE = "/home/user/Downloads.internal/CB_MASTER_live.xlsx"
SRC = "/home/user/Downloads.internal/Items Identifed During Update Meeting 051926 Status 0605 UPDATED Rich Format MATCHED CB - 0611 SCENARIOS.xlsx"

def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())[:60]

# per-ticket scenario finishes: (S1 all-parallel, S2 focused REC, S3 DPP-only)
TIMING = {
    "SF-3333": ("Fri Jun 19", "Wed Jun 17", "Tue Jun 16"),
    "SF-3329": ("Fri Jun 26", "Fri Jun 26", "Wed Jun 24"),
    "SF-3330": ("Thu Jul 2", "Wed Jul 1", "Mon Jun 29"),
    "SF-3332": ("Fri Jul 3", "Thu Jul 2", "Tue Jun 30"),
    "SF-3334": ("Wed Jul 8", "Fri Jun 19", "Thu Jun 18"),
    "SF-3335": ("Fri Jul 10", "Fri Jun 26", "Wed Jun 24"),
}
PARKED_TIMING = "Parked - not scheduled this cycle (any scenario)"
DECISION_TIMING = "Blocked on decision - not schedulable in any scenario until decided"

# extra rows not in the 0611 SCENARIOS main tab (topic-prefix -> status, comment-append, effort, delivery)
PATCH = {
    "doweorwillwehavethenecessaryforms": ("PARKED - NEEDS TICKET",
        "Sign-off forms need a ticket; closest neighbor DPP-019 (SF-3347), parked - est 5-8d next cycle", 0, "Not scheduled"),
    "q10needtoproducethehistoricalpartnership": ("PARKED - AWAITING DECISION",
        "DPP-017 (SF-3345) discovery DONE (727 agreements, 489 tie to dealer, no dealer field on envelopes); loading blocked on Iliane/data-owner call", 0, "Not scheduled"),
    "q19timelywarrantclaims": ("PARKED - AWAITING DECISION",
        "DPP-022 (SF-3350) pure business-rule call (grace period, points timing; Satish K = Tavant contact); no build until rule written", 0, "Not scheduled"),
    "needlistofminimumrequirementsinacolumnar": ("PARKED - QUICK-WIN CANDIDATE",
        "DPP-014 (SF-3342) layout quick win ~1d; slot-in any gap week, off critical path", 0, "Slot-in (any gap week)"),
    "forthesafetymeetinginputwhenthedealeradds": ("PARKED - NEXT CYCLE",
        "UX enhancement; rides min-req checklist wave (DPP-020/SF-3348) next cycle", 0, "Not scheduled"),
    "howdowegetdealerstoatleastannuallyupdate": ("PARKED - AWAITING DECISION",
        "Program-scope decision first (min-req vs separate process); no ticket", 0, "Not scheduled"),
    "dontlockuploadsetcuntilevaluationiscompleted": ("DECIDED - IN SCOPE (config)",
        "Jun-8 meeting DECIDED: Year-1 stays flexible, lock only after TM approval/eval; config rides Wave 2-3 validation", 0.5, "By Jul 2 (rides Wave 2-3)"),
    "canthedashboardcontainthebulletinandwebinar": ("PARKED - QUICK-WIN CANDIDATE",
        "DPP-014 (SF-3342) ~0.5-1d; explicit Jun-8 next-step (bulletin/webinar component); asset link with Operator", 0, "Slot-in (any gap week)"),
    "queanemailonlyifyouhaveitemstoreview": ("PARKED - AWAITING DECISION",
        "Quiet-email spec drafted in DPP-011 (SF-3339); parked with approval wave pending Steve's DPP-010 status sign-off", 0, "Not scheduled"),
    "anapprovaldashboardseeallpertm": ("PARKED - AWAITING DECISION",
        "DPP-011 (SF-3339) TM queue spec drafted; blocked on DPP-010 status vocabulary; 10-15d approval wave counted once", 0, "Not scheduled"),
    "abilitytohaveapprovalsbydifferentusers": ("NEXT YEAR-PARKING LOT",
        "Future-state per Rich; design single-approver flow first", 0, "Not scheduled"),
    "pmpnextyear": (None, None, 0, "Not scheduled"),
    "evaluateq10purchasepartnership": (None, None, 0, "Not scheduled"),
    "discussionitemprogramforfinanicialrewards": (None, None, 0, "Not scheduled"),
    "onpage8ofmarketshareforminhandbookthefirsttwolines": (None, None, 0, "Not scheduled"),
}

src = openpyxl.load_workbook(SRC)
smap = {}
for row in src["Topics Identified & Status"].iter_rows(min_row=2, values_only=True):
    topic, comments, status, priority, effort, delivery = row[:6]
    if topic:
        smap[norm(topic)] = {"comments": comments, "status": status, "priority": priority,
                             "effort": effort, "delivery": delivery}

def lookup(topic):
    k = norm(topic)
    if k in smap:
        return dict(smap[k]), None
    for sk, v in smap.items():
        if sk.startswith(k[:42]) or k.startswith(sk[:42]):
            return dict(v), None
    for pk, (st, app, eff, dlv) in PATCH.items():
        if k.startswith(pk[:40]) or pk.startswith(k[:30]):
            return None, (st, app, eff, dlv)
    return None, None

def timing_for(comment, status):
    s = (status or "").upper()
    if "PARKED" in s or "NEXT YEAR" in s or "BLOCKED" in s or "NOT SCHEDULED" in (comment or "").upper():
        return PARKED_TIMING
    if "DECISION" in s:
        return DECISION_TIMING
    m = re.search(r"SF-33(29|30|31|32|33|34|35)", comment or "")
    if m:
        t = TIMING.get("SF-33" + m.group(1))
        if t:
            return f"Scen 1 (all parallel): {t[0]}; Scen 2 (focused, REC): {t[1]}; Scen 3 (DPP-only): {t[2]}"
    return PARKED_TIMING

wb = openpyxl.load_workbook(LIVE)

# ---- Tab 1: Topics Identified & Status (A topic, B category, C comments, D status, E pri, F effort, G timing)
ws = wb["Topics Identified & Status"]
t1 = 0
for r in range(2, ws.max_row + 1):
    topic = ws.cell(row=r, column=1).value
    if not topic or not str(topic).strip():
        continue
    hit, patch = lookup(str(topic))
    if hit:
        ws.cell(row=r, column=3).value = hit["comments"]
        ws.cell(row=r, column=4).value = hit["status"]
        ws.cell(row=r, column=5).value = hit["priority"]
        ws.cell(row=r, column=6).value = hit["effort"]
        ws.cell(row=r, column=7).value = timing_for(hit["comments"], str(hit["status"])) if "PARKED" not in str(hit["status"]).upper() and "DECISION" not in str(hit["status"]).upper() else (DECISION_TIMING if "DECISION" in str(hit["status"]).upper() else PARKED_TIMING)
        # active rows: timing from ticket in comments
        if "IN PROGRESS" in str(hit["status"]).upper():
            ws.cell(row=r, column=7).value = timing_for(hit["comments"], "")
        t1 += 1
    elif patch:
        st, app, eff, dlv = patch
        if st:
            ws.cell(row=r, column=4).value = st
        if app:
            c = ws.cell(row=r, column=3)
            c.value = f"{c.value or ''}  ||  {app}"
        ws.cell(row=r, column=5).value = 3
        ws.cell(row=r, column=6).value = eff
        ws.cell(row=r, column=7).value = ("By Jul 2 - rides Wave 2-3 validation (all scenarios)" if eff else PARKED_TIMING)
        t1 += 1

# ---- Tab 2: Estimated_List (A topic, B comments, C status, D pri, E effort, F delivery)
ws2 = wb["Estimated_List"]
t2 = 0
for r in range(2, ws2.max_row + 1):
    topic = ws2.cell(row=r, column=1).value
    if not topic or not str(topic).strip() or "Total" in str(ws2.cell(row=r, column=4).value or ""):
        continue
    hit, patch = lookup(str(topic))
    if hit:
        ws2.cell(row=r, column=2).value = hit["comments"]
        ws2.cell(row=r, column=3).value = hit["status"]
        ws2.cell(row=r, column=4).value = hit["priority"]
        ws2.cell(row=r, column=5).value = hit["effort"]
        ws2.cell(row=r, column=6).value = hit["delivery"]
        t2 += 1
    elif patch:
        st, app, eff, dlv = patch
        if st:
            ws2.cell(row=r, column=3).value = st
        if app:
            c = ws2.cell(row=r, column=2)
            c.value = f"{c.value or ''}  ||  {app}"
        ws2.cell(row=r, column=4).value = 3
        ws2.cell(row=r, column=5).value = eff
        ws2.cell(row=r, column=6).value = dlv
        t2 += 1
# notes under the total
ws2.cell(row=48, column=4).value = "(+ SF-3331 mapping integrity 3-4d - no topic row; 4,352/4,357 records clean, 5 orphan confirmations then backfill; gates multi-location rows)"
ws2.cell(row=49, column=4).value = "DPP critical-path total incl SF-3331 = 27-32 person-days (~216-256 hrs). SPP MVP parallel (SF-3128 8-12d, SF-2490 2-4d). Scen 2 (REC) finish Jun 30-Jul 3."

# ---- Tab 3: Work_Items_Out (no header; A topic, B comments, C status, D pri, E effort, F ?, G formula)
ws3 = wb["Work_Items_Out"]
t3 = 0
for r in range(1, ws3.max_row + 1):
    topic = ws3.cell(row=r, column=1).value
    if not topic or not str(topic).strip():
        continue
    hit, patch = lookup(str(topic))
    if hit:
        ws3.cell(row=r, column=2).value = hit["comments"]
        ws3.cell(row=r, column=3).value = hit["status"]
        ws3.cell(row=r, column=4).value = hit["priority"]
        ws3.cell(row=r, column=5).value = hit["effort"]
        t3 += 1
    elif patch:
        st, app, eff, dlv = patch
        if st:
            ws3.cell(row=r, column=3).value = st
        if app:
            c = ws3.cell(row=r, column=2)
            c.value = f"{c.value or ''}  ||  {app}"
        ws3.cell(row=r, column=4).value = 3
        ws3.cell(row=r, column=5).value = eff
        t3 += 1

wb.save(LIVE)
print(f"updated: tab1={t1} rows, tab2={t2} rows, tab3={t3} rows -> {LIVE}")
