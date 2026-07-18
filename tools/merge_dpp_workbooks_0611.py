#!/usr/bin/env python3
"""Merge the 0611 SCENARIOS content into the 0605 master workbook layout (master wins on structure)."""
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

M = "/home/user/Downloads.internal/Items Identifed during Update Meeting 051926 - status 0605.xlsx"
S = "/home/user/Downloads.internal/Items Identifed During Update Meeting 051926 Status 0605 UPDATED Rich Format MATCHED CB - 0611 SCENARIOS.xlsx"
OUT = "/home/user/Downloads.internal/Items Identifed during Update Meeting 051926 - status 0611 MERGED.xlsx"

def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())[:70]

src = openpyxl.load_workbook(S)
st = src["Topics Identified & Status"]
smap = {}
for row in st.iter_rows(min_row=2, values_only=True):
    topic, comments, status, priority, effort, delivery = row[:6]
    if topic:
        smap[norm(topic)] = (comments, status, priority, effort, delivery)

wb = openpyxl.load_workbook(M)
ws = wb["Topics Identified & Status"]
matched, unmatched_master = 0, []
used = set()
for r in range(3, ws.max_row + 1):
    topic = ws.cell(row=r, column=1).value
    if not topic or not str(topic).strip():
        continue
    k = norm(str(topic))
    hit = None
    if k in smap:
        hit = smap[k]; used.add(k)
    else:
        for sk in smap:
            if sk.startswith(k[:45]) or k.startswith(sk[:45]):
                hit = smap[sk]; used.add(sk); break
    if hit:
        comments, status, priority, effort, delivery = hit
        ws.cell(row=r, column=3).value = status
        eff = f"{effort:g}d" if isinstance(effort, (int, float)) else str(effort)
        ws.cell(row=r, column=4).value = (
            f"{comments}  ||  PRIORITY {priority} | EFFORT {eff} | EST DELIVERY: {delivery}"
        )
        matched += 1
    else:
        unmatched_master.append(str(topic)[:70])

wrap = Alignment(wrap_text=True, vertical="top")
bold = Font(bold=True)
hdr_fill = PatternFill("solid", fgColor="DDEBF7")
for name in ("DPP Reprioritized Craig Model", "Executive Reprioritized"):
    sws = src[name]
    nws = wb.create_sheet(name)
    for row in sws.iter_rows(values_only=True):
        nws.append(list(row))
    for row in nws.iter_rows():
        for c in row:
            c.alignment = wrap
    for c in nws[1]:
        c.font = bold
        if name.startswith("DPP"):
            c.fill = hdr_fill
    for col, dim in sws.column_dimensions.items():
        if dim.width:
            nws.column_dimensions[col].width = dim.width

wb.save(OUT)
print("SAVED:", OUT)
print(f"matched {matched} master rows; 0611 rows not consumed: {len(smap)-len(used)}")
for sk, v in smap.items():
    if sk not in used:
        print("  UNUSED 0611:", sk[:60])
print("master rows w/o 0611 update:", len(unmatched_master))
for t in unmatched_master:
    print("  KEPT AS-IS:", t)
