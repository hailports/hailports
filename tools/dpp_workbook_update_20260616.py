#!/usr/bin/env python3
"""One-shot: reflect 6/16 DPP work into the master workbook.
Tab 1 (Topics Identified & Status): update Status + append dated note to touched rows.
Tab 2 (Estimated_List): update Remaining Hours (G) + Status + Owner Notes IN PLACE (no new cols).
Backs up first. Owner-authorized (Operator's own OneDrive file)."""
import openpyxl, shutil, time, os, zipfile

SRC = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-redactedIndustries,Inc/MASTER_DPP_SPP_FULLY_UPDATED_FINAL.xlsx")
BK_DIR = os.path.expanduser("~/claude-stack/data/dpp/backups")
os.makedirs(BK_DIR, exist_ok=True)
stamp = time.strftime("%Y%m%d-%H%M%S")
bk = os.path.join(BK_DIR, f"MASTER_DPP_SPP_{stamp}.xlsx")
shutil.copy2(SRC, bk)
print("backup:", bk)

# chart/image presence check (openpyxl may drop these on save)
with zipfile.ZipFile(SRC) as z:
    names = z.namelist()
charts = [n for n in names if "/charts/" in n or n.endswith(".emf")]
imgs   = [n for n in names if "/media/" in n]
pivots = [n for n in names if "pivotTable" in n or "pivotCache" in n]
print(f"pre-save assets: charts={len(charts)} images={len(imgs)} pivotParts={len(pivots)}")

NOTE = "  •[6/16] "

# ---- TAB 1: row -> (new_status_or_None, append_note) ----
T1 = {
 3:  ("BUILT — access SHIPPED; piloting login-as",
      "TM + dealer submission-item edit and TM approve SHIPPED to prod (login-as inherits dealer perms). E-commerce-hide on login-as still pending Craig Bell sign-off. SF-3351 -> Lucas."),
 5:  (None,
      "Rich confirmed tiers 90/80/70/60 + handbook alignment; 'Valley Dealer' Dealer_Type value shipped to prod. Ranking formula still blocked (formula compile-size + DocuSign match) -> SF-3334 kept joint Operator+Cinthya."),
 6:  (None,
      "SHIPPED to prod: Brazil dashboard-folder leak closed (129 partner shares -> 1 internal). DppActivityTask isolation + sharing set staged in PARTIAL; needs manual Sharing-Settings recalc -> re-verify -> prod. SF-3330 joint Operator+Cinthya."),
 8:  (None,
      "Heidi confirmed 3-yr-avg columns + market-share window Aug 1 - Jul 31; analyst validation continues. SF-3344 joint Operator+Cinthya."),
 11: (None,
      "SF-3330 portion: dashboard-folder leak fixed in prod; record isolation staged in partial (recalc pending)."),
 12: (None,
      "SF-3334 tiers confirmed (Rich 90/80/70/60)."),
 14: (None,
      "SF-3330: folder leak fixed in prod; isolation staged (recalc pending). Joint Operator+Cinthya."),
 17: ("DECISION CONFIRMED (Steve) — build pending Databricks",
      "Steve confirmed net-connection definition (CONN-DEF-11: count dealer-linked only, exclude both transfer flags + received-out, include received-from-same-dealer-location). Recipe lives in Databricks (Ravi's team); goal-display shell rides SF-3344."),
 18: (None,
      "Steve net-conn definition confirmed (CONN-DEF-11); goal-display shell ships, number pends Databricks feed."),
 23: ("IN PROGRESS — PARTIAL SHIPPED",
      "SHIPPED to prod: Nick's dealer 'My Account' field lockdown (read-only vs dealer-edit FLS + community layout) + 5 Account social-URL fields (FB/X/LinkedIn/IG/YT). Footprint URL->Account auto-sync flow built + staged in PARTIAL (SF-3386 -> Cinthya; needs change-detection gate before promote). SF-3347 -> Bryan."),
 25: ("ACCESS FIXED — report rebuild pending",
      "Dashboard/report-type access fixed in prod (red errors cleared). Full Q18 report re-plumb still pending under SF-3346."),
 26: (None,
      "On hold — blocked on Satish's Tavant cutover date. SF-3350 joint Operator+Cinthya."),
 27: (None,
      "SF-3330 prod fixes landed (submission-tab data access + folder leak); UAT validation continues."),
 30: (None,
      "Field/dropdown built (PR #288: MR_BusinessSuccessionPlan__c picklist Yes-Plan-on-File / In Progress). Checklist UI wave stays parked next cycle. SF-3348 -> Cinthya."),
 38: (None,
      "Min-req field built (PR #288: MR_CertificateLiabilityInsurance__c + MR_FinancialInformationOnfile__c + FLS). Matt Person / Mark Kable process blocker remains. SF-3348 -> Cinthya."),
 39: (None,
      "Min-req field built (PR #288: MR_MaintainFinancialAccounts__c + FLS). Same Matt/Mark blocker. SF-3348 -> Cinthya."),
 42: (None,
      "SF-3334 tiers confirmed (Rich); Q11 number pends Steve's net-conn feed (Databricks)."),
 43: (None,
      "SF-3330 folder leak fixed in prod; isolation staged (recalc pending)."),
 54: (None,
      "SF-3334 tiers + per-Q max-points confirmed (Rich)."),
 55: (None,
      "SF-3344: Heidi confirmed market-share column window (Aug 1 - Jul 31; 23/24, 24/25, 25/26)."),
 57: (None,
      "SF-3344: Heidi confirmed EDA 3-yr-avg + by-territory; analyst validation continues."),
 64: (None,
      "SF-3330: folder leak fixed in prod; isolation staged (recalc pending)."),
 65: (None,
      "SF-3344/SF-3334: Heidi confirmed Y1/Y2/Y3 + 3-yr-avg display source; tiers confirmed (Rich)."),
}

# ---- TAB 2 (Estimated_List): row -> (new_status_or_None, new_remaining_hours_or_None, owner_note_append_or_None) ----
# cols: C=status(3) G=remaining hrs(7) M=owner notes(13)
T2 = {
 2:  (None, 1,   "6/16: folder leak fixed in prod; isolation + sharing-set staged in partial, manual recalc pending."),
 7:  (None, 0.5, "6/16: same fix as Q2 (sharing flip); folder leak already in prod."),
 8:  (None, 5,   "6/16: report/dashboard access fixed in prod, eases UAT; validation continues."),
 9:  (None, 0.5, "6/16: SF-3330 folder leak fixed in prod; isolation staged."),
 15: (None, 5,   "6/16: tiers confirmed (Rich) + 'Valley Dealer' value shipped; remaining = ranking formula design + DocuSign match."),
 19: ("ACCESS FIXED — rebuild pending", 20, "6/16: report-type access fixed in prod (red errors cleared); report re-plumb remains."),
 20: (None, 3,   "6/16: min-req fields built (PR #288); checklist UI wave still parked."),
 21: (None, 0,   "6/16: field/dropdown built + deployed (PR #288)."),
 28: (None, 2.5, "6/16: cert/financial fields built (PR #288); Matt/Mark process blocker remains."),
 29: (None, 0.25,"6/16: MR_MaintainFinancialAccounts__c built (PR #288)."),
}

wb = openpyxl.load_workbook(SRC)  # keep formulas/formatting
ws1 = wb["Topics Identified & Status"]
ws2 = wb["Estimated_List"]

n1 = 0
for r,(st,note) in T1.items():
    if st is not None:
        ws1.cell(row=r, column=3).value = st
    cur = ws1.cell(row=r, column=4).value or ""
    if "[6/16]" not in str(cur):
        ws1.cell(row=r, column=4).value = str(cur) + NOTE + note
    n1 += 1

n2 = 0
for r,(st,rem,note) in T2.items():
    if st is not None:
        ws2.cell(row=r, column=3).value = st
    if rem is not None:
        ws2.cell(row=r, column=7).value = rem
    if note:
        cur = ws2.cell(row=r, column=13).value or ""
        if "6/16:" not in str(cur):
            ws2.cell(row=r, column=13).value = str(cur) + " | " + note
    n2 += 1

wb.save(SRC)
print(f"updated tab1 rows={n1}, tab2 rows={n2}")

# verify charts/images survived
with zipfile.ZipFile(SRC) as z:
    names2 = z.namelist()
c2 = len([n for n in names2 if "/charts/" in n or n.endswith(".emf")])
i2 = len([n for n in names2 if "/media/" in n])
p2 = len([n for n in names2 if "pivotTable" in n or "pivotCache" in n])
print(f"post-save assets: charts={c2} images={i2} pivotParts={p2}")
if (c2,i2,p2) != (len(charts),len(imgs),len(pivots)):
    print("WARN: asset count changed on save — restore from backup if formatting broke:", bk)
